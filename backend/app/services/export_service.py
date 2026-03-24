"""
Export utility service.

Generates CSV and Excel files from query result data,
returning FastAPI StreamingResponse objects for download.
"""

import csv
import io
from typing import Any, Sequence

from fastapi.responses import StreamingResponse


def export_to_csv(
    data: Sequence[dict[str, Any]],
    columns: list[str] | None = None,
    filename: str = "export.csv",
) -> StreamingResponse:
    """
    Generate a CSV file from a list of dicts and return as a streaming download.

    Args:
        data: List of row dicts (e.g., from SQLAlchemy .mappings()).
        columns: Ordered list of column names to include. If None, uses keys from first row.
        filename: Name for the downloaded file.

    Returns:
        FastAPI StreamingResponse with CSV content.
    """
    if not data:
        # Return empty CSV with just headers
        columns = columns or []
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(columns)
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if columns is None:
        columns = list(data[0].keys())

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()

    for row in data:
        # Ensure all values are serializable
        clean_row = {}
        for col in columns:
            val = row.get(col)
            if val is None:
                clean_row[col] = ""
            elif isinstance(val, (list, dict)):
                import json
                clean_row[col] = json.dumps(val)
            else:
                clean_row[col] = str(val)
        writer.writerow(clean_row)

    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def export_to_excel(
    data: Sequence[dict[str, Any]],
    columns: list[str] | None = None,
    sheet_name: str = "Sheet1",
    filename: str = "export.xlsx",
) -> StreamingResponse:
    """
    Generate an Excel (.xlsx) file from a list of dicts and return as a streaming download.

    Args:
        data: List of row dicts.
        columns: Ordered list of column names to include. If None, uses keys from first row.
        sheet_name: Name for the Excel worksheet.
        filename: Name for the downloaded file.

    Returns:
        FastAPI StreamingResponse with Excel content.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        columns = columns or []
    elif columns is None:
        columns = list(data[0].keys())

    # Style the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for row_idx, row in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = row.get(col_name)
            if isinstance(val, (list, dict)):
                import json
                val = json.dumps(val)
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-adjust column widths (approximate)
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, min(len(data) + 2, 52)):  # sample first 50 rows
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    # Write to bytes buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
